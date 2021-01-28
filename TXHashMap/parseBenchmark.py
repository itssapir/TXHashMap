import openpyxl
from openpyxl.utils.cell import get_column_letter

def parse(file):
    f = open(file, 'r')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = "Len \\ Num Jobs"
    col = -2
    prevNumJobs = -1
    row = 2
    while True:
        row = row + 1
        head = f.readline()
        txPut = f.readline().split(" ")[-1][0:-1]
        txGet = f.readline().split(" ")[-1][0:-1]
        oraclePut = f.readline().split(" ")[-1][0:-1]
        oracleGet = f.readline().split(" ")[-1][0:-1]
        if not head:
            break
        numJobs = head.split(" ")[4]
        numPerJob = head.split(" ")[8]
        warmupCycles = head.split(" ")[12][0:-1]
        
        if numJobs != prevNumJobs:
            prevNumJobs = numJobs
            row = 3
            col = col + 4
            txColPut = get_column_letter(col)
            txColGet = get_column_letter(col + 1)
            orColPut = get_column_letter(col + 2)
            orColGet = get_column_letter(col + 3)
            sheet[txColPut+'1'] = str(numJobs)
            sheet[txColPut+'2'] = 'TX - Put'
            sheet[orColPut+'2'] = 'Oracle - Put'
            sheet[txColGet+'2'] = 'TX - Get'
            sheet[orColGet+'2'] = 'Oracle - Get'

        sheet['A'+ str(row)] = numPerJob
        sheet[txColPut + str(row)] = int(txPut)
        sheet[txColGet + str(row)] = int(txGet)
        sheet[orColPut + str(row)] = int(oraclePut)
        sheet[orColGet + str(row)] = int(oracleGet)
        
    wb.save(filename = file + ".xlsx")



if __name__ == "__main__":
    parse("res.txt")